from fastapi import FastAPI
from pydantic import BaseModel
from typing import List, Dict, Any
import pandas as pd
import os
import uuid
import logging
import re
import openpyxl

from dropbox_uploader import upload_to_dropbox

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI()

class RequestModel(BaseModel):
    fileName: str
    data: List[Dict[str, Any]]

def sanitize_filename(filename: str) -> str:
    """Rimuove i caratteri non alfanumerici, lasciando solo lettere, numeri, -, _, ."""
    return re.sub(r"[^\w\-\.]", "", filename)

def short_uid(length=5) -> str:
    """Genera un ID univoco corto"""
    return uuid.uuid4().hex[:length]

@app.post("/upload-json/")
async def upload_json(payload: RequestModel):
    logger.info("📥 Ricevuta richiesta JSON per conversione in XLSX.")
    filepath = None

    try:
        # 1. Nome sicuro + ID corto
        safe_name = sanitize_filename(payload.fileName)
        unique_id = short_uid()
        final_filename = f"{safe_name}_{unique_id}.xlsx"
        filepath = os.path.join("/tmp", final_filename)
        logger.info(f"📄 Nome file generato: {final_filename}")

        # 2. Converti a DataFrame
        df = pd.DataFrame(payload.data)

        # 3. Pulizia e conversione numerica, esclusi 'BC' e 'IN'
        for col in df.columns:
            if col in ["BC", "IN"]:
                df[col] = df[col].astype(str).str.strip()
                continue

            if df[col].dtype == object:
                df[col] = df[col].astype(str)
                df[col] = df[col].str.replace("€", "", regex=False)
                df[col] = df[col].str.replace(",", ".", regex=False)
                df[col] = df[col].str.strip()

                try:
                    df[col] = pd.to_numeric(df[col])
                except Exception:
                    pass

        logger.info(f"🧾 Colonne: {df.columns.tolist()}")

        # 4. Salvataggio Excel
        df.to_excel(filepath, index=False)

        # 5. Formattazione numerica Excel (2 decimali) + forzatura testo su BC e IN
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'

        # Forza BC e IN come testo (number_format = '@')
        for col in ["BC", "IN"]:
            if col in headers:
                col_letter = openpyxl.utils.get_column_letter(headers[col])
                for cell in ws[col_letter][1:]:  # skip header
                    cell.number_format = '@'

        wb.save(filepath)

        logger.info(f"📄 File Excel salvato: {filepath}")

        # 6. Upload su Dropbox in /mtoXLSX
        dropbox_path = f"/mtoXLSX/{final_filename}"
        response = upload_to_dropbox(filepath, dropbox_path)
        logger.info(f"✅ Caricato su Dropbox: {dropbox_path}")

        return {
            "message": "Upload completato",
            "dropbox_path": dropbox_path,
            "dropbox_response": response
        }

    except Exception as e:
        logger.error(f"❌ Errore: {str(e)}")
        return {"error": str(e)}

    finally:
        if filepath and os.path.exists(filepath):
            os.remove(filepath)
            logger.info(f"🧹 File temporaneo eliminato: {filepath}")
